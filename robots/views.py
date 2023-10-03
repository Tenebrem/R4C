from datetime import datetime, timedelta

from django.db.models import Count
from django.http import HttpResponse
from django.views.generic import View
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.styles.alignment import Alignment

from .models import Robot

DAY_PERIOD = 7
COLUMN_WIDTH = 22


class ExcelExportView(View):
    """ Создает excele файл с данным за неделю."""
    def get(self, request):
        # Узнаем начало и конец прошлой недели
        today = datetime.now()
        days_to_subtract = today.weekday() + DAY_PERIOD
        start_of_last_week = today - timedelta(days=days_to_subtract)
        end_of_last_week = start_of_last_week + timedelta(days=DAY_PERIOD-1)
        # Запращиваем в бд даныне за этот период
        data = Robot.objects.filter(
            created__date__range=(
                start_of_last_week.date(),
                end_of_last_week.date()
            )
        )
        # Считаем похожие версии
        count_by_version_model = data.values(
            'version', 'model'
            ).annotate(
                count=Count('id')
            )
        # Создаем таблицы в excele
        workbook = Workbook()
        border_style = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

        for robot in count_by_version_model:
            model = robot['model']
            version = robot['version']
            count = robot['count']

            if model not in workbook.sheetnames:
                worksheet = workbook.create_sheet(model)
                worksheet.append(['Модель', 'Версия', 'Количество за неделю'])
                worksheet.column_dimensions['C'].width = COLUMN_WIDTH
            else:
                worksheet = workbook[model]

            worksheet.append([model, version, count])
        workbook.remove(workbook['Sheet'])
        # делаем заливку границ и выравниваем текст
        for sheet in workbook.sheetnames:
            current_sheet = workbook[sheet]
            for row in current_sheet:
                for cell in row:
                    cell.border = border_style
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')

        response = HttpResponse(
            content_type='application/vnd.'
            'openxmlformats-officedocument.spreadsheetml.sheet'
            )
        response['Content-Disposition'] = ('attachment;'
                                           'filename=robot_production.xlsx'
                                           )
        workbook.save(response)
        return response
